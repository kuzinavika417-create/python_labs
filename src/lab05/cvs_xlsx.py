import csv
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """
    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f'CSV файл не найден')
    if not csv_path.lower().endswith('.csv'):
        raise ValueError(f'Файл должен иметь расширение csv')
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            line = f.readline()
            if not line.strip():
                raise ValueError('Пустой CSV файл')
            f.seek(0)  # Возвращаемся в начало файла
            reader = csv.reader(f)
            data = list(reader)
    except UnicodeDecodeError:
        raise ValueError(f'неправильная кодировка (должна быть UTF-8)')
    except csv.Error:
        raise ValueError(f'неправильный формат CSV файла')
    except Exception as ex:
        raise ValueError(f'Ошибка при чтении файла CSV: {ex}')
    
    if len(data) == 0:
        raise ValueError("CSV файл не содержит данных")
    if len(data) < 1:
        raise ValueError("CSV файл не содержит заголовки")
    
    try:
        work = openpyxl.Workbook()
        sheet = work.active
        title = 'Sheet1'
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        for col_idx in range(1, len(data[0]) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 8  # минимальная ширина колонок
            for row in sheet[column_letter]:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            sheet.column_dimensions[column_letter].width = max_length + 2 # добавление отступов
        work.save(xlsx_path)
    except Exception as ex:
        raise ValueError(f"Ошибка при создании XLSX файла: {ex}")